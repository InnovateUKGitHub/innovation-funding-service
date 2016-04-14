from openpyxl import load_workbook

__author__ = 'Ika Belerma'


class XlsxLibrary:

    ROBOT_LIBRARY_SCOPE = 'TEST SUITE'

    def __init__(self):
        self.data = []

    def open_excel_file(self, path_to_excel_file):
        """Returns a workbook instance from the an xls file provided.
        Other keywords will ask for an instantiated workbook as an argument.
        Example:
        | ${workbook} | Open Excel File | Book1.xls |
        """
        workbook = load_workbook(path_to_excel_file)
        return workbook

    def get_cell_value_by_sheet_name(self, workbook, sheet_name, cell_name):
        """Returns the value of the cell indicated, given the sheet name.
        Example:
        | ${workbook} | Open Excel File | Book1.xls |
        | ${A1} | Get Cell Value By Sheet Name | ${workbook} | Sheet1 | A1 |
        This example returns the value of cell "A1" from "Sheet 1" from the xls file "Book1.xls".
        """
        workbook_sheet = workbook.get_sheet_by_name(sheet_name)
        cell_value = self._return_cell_value(workbook_sheet, cell_name)
        return cell_value

    def get_cell_value_of_active_sheet(self, workbook, cell_name):
        """Returns the value of the cell indicated of the currently active sheet.
        Example:
        | ${workbook} | Open Excel File | Book1.xls |
        | ${A1} | Get Cell Value Of Active Sheet | ${workbook} | A1 |
        This example returns the value of cell "A1" from "Sheet 1" from the xls file "Book1.xls".
        """
        workbook_sheet = workbook.get_active_sheet()
        cell_value = self._return_cell_value(workbook_sheet, cell_name)
        return cell_value

    def get_sheet_names(self, workbook):
        """Returns the names of the sheets of the workbook provided.
        Example:
        | ${workbook} | Open Excel File | Book1.xls |
        | ${sheet_names} | Get Sheet Names | ${workbook} |
        Given Book1.xls has three sheets with names Sheet1, Sheet2 and Sheet3:
        
        | Log | ${sheet_names} | # ['Sheet1', 'Sheet2', 'Sheet3'] |
        | Log | ${sheet_names[0]} | # Sheet1 |
        | Log | ${sheet_names[1]} | # Sheet2 |
        | Log | ${sheet_names[2]} | # Sheet3 |
        """
        sheet_names = workbook.get_sheet_names()
        return sheet_names

    @staticmethod
    def _return_cell_value(workbook_sheet, cell_name):
        cell_value = workbook_sheet.cell(cell_name).value
        return cell_value
